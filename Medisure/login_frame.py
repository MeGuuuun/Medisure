import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
from bcrypt_utils import check_password
from PIL import Image, ImageTk

EXCEL_PATH = "USER_DOCS.xlsx"
IMAGE_PATH = "Medisure_logo.png"

# 유저 정보 불러오는 함수
def load_credentials(path):
    credentials = {}
    try:
        wb = load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:  # ID와 비밀번호 모두 있을 때만
                credentials[row[0]] = row[1]
        wb.close()
    except FileNotFoundError:
        print(f"⚠️ 경고: '{path}' 파일을 찾을 수 없습니다.")
    except Exception as e:
        print(f"⚠️ 엑셀 파일 로드 중 오류 발생: {e}")
    return credentials

# 로그인 프레임 생성
def create_login_frame(root, switch_to_join, on_login_success):
    frame = tk.Frame(root)

    # 로고
    logo_frame = tk.Frame(frame, width=150, height=150)
    logo_frame.pack(pady=(30, 10))
    logo_frame.pack_propagate(False)

    logo_img_raw = Image.open(IMAGE_PATH)
    logo_img_resized = logo_img_raw.resize((150, 150), Image.Resampling.LANCZOS)
    logo_img = ImageTk.PhotoImage(logo_img_resized)

    logo_label = tk.Label(logo_frame, image=logo_img)
    logo_label.image = logo_img
    logo_label.pack()

    tk.Label(frame, text="로그인", font=("Arial", 22, "bold")).pack(pady=10)

    # ID
    tk.Label(frame, text="ID", font=("Arial", 14)).pack(pady=(10, 2))
    entry_id = tk.Entry(frame, font=("Arial", 12), width=25)
    entry_id.pack(ipady=5)

    # Password
    tk.Label(frame, text="Password", font=("Arial", 14)).pack(pady=(20, 2))
    entry_pw = tk.Entry(frame, show="*", font=("Arial", 12), width=25)
    entry_pw.pack(ipady=5)

    tk.Frame(frame, height=30).pack()

    # 로그인 처리
    def try_login():
        user_id = entry_id.get().strip()
        password = entry_pw.get().strip()

        if not user_id:
            messagebox.showwarning("입력 오류", "ID를 입력하세요.")
            return
        if not password:
            messagebox.showwarning("입력 오류", "비밀번호를 입력하세요.")
            return

        credentials = load_credentials(EXCEL_PATH)
        if not credentials:
            messagebox.showerror("시스템 오류", "사용자 데이터 로드에 실패했습니다.")
            return

        try:
            if user_id not in credentials:
                messagebox.showinfo("로그인 실패", "존재하지 않는 아이디입니다.")
            elif not check_password(password, credentials[user_id]):
                messagebox.showwarning("로그인 실패", "비밀번호가 틀렸습니다.")
            else:
                messagebox.showinfo("로그인 성공", f"{user_id}님 환영합니다!")
                on_login_success(user_id)
        except Exception as e:
            messagebox.showerror("오류", f"로그인 처리 중 오류가 발생했습니다:\n{e}")

    button_frame = tk.Frame(frame)
    button_frame.pack(pady=(30, 10))

    tk.Button(button_frame, text="로그인", font=("Arial", 13), width=15, command=try_login).pack(side="left",padx=5)
    tk.Button(button_frame, text="회원가입", font=("Arial", 13),width=15, command=switch_to_join).pack(side="left", padx=5)

    return frame
