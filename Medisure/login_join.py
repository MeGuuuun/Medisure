import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
import bcrypt

EXCEL_PATH = "USER_DOCS.xlsx"

# ===== 유틸 함수 =====

def hash_password(plain_password):
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(plain_password.encode(), salt).decode()

def check_password(plain_password, hashed_password):
    return bcrypt.checkpw(plain_password.encode(), hashed_password.encode())

def load_credentials(path):
    credentials = {}
    try:
        wb = load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            credentials[row[0]] = row[1]
    except FileNotFoundError:
        pass
    return credentials

# 창을 모니터 정중앙에 위치 + 창 크기 조절 불가 + 로고 위치 고정
def default_window(root, width, height):
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    center_x = int((screen_width - width) / 2)
    center_y = int((screen_height - height) / 2)

    root.geometry(f"{width}x{height}+{center_x}+{center_y}")
    root.resizable(False, False)

    root.attributes("-topmost", True)

    # 로고 부분
    logo_frame = tk.Frame(root, width=150, height=150, bg="lightblue")
    logo_frame.pack(pady=(100,50))
    logo_frame.pack_propagate(False)

# ===== 회원 가입 창 =====

def open_join_window(login_frame):
    login_frame.pack_forget()

    join_window = tk.Toplevel(root)
    join_window.title("회원가입")

    default_window(join_window, 450, 650)

    # 회원가입을 하지 않고 창을 닫는 경우
    def on_close():
        join_window.destroy()
        login_frame.pack(fill="both", expand=True)

    join_window.protocol("WM_DELETE_WINDOW", on_close)

    tk.Label(join_window, text="회원가입", font=("Arial", 16)).pack(pady=10)

    tk.Label(join_window, text="ID").pack()
    entry_id = tk.Entry(join_window)
    entry_id.pack()

    tk.Label(join_window, text="Password").pack()
    entry_pw = tk.Entry(join_window, show="*")
    entry_pw.pack()

    def save_user():
        user_id = entry_id.get().strip()
        password = entry_pw.get().strip()

        if not user_id or not password:
            messagebox.showwarning("입력 오류", "ID와 Password 모두 입력하세요.")
            return

        credentials = load_credentials(EXCEL_PATH)

        if user_id in credentials:
            messagebox.showerror("중복 오류", "이미 존재하는 ID입니다.")
            return

        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        ws.append([user_id, hash_password(password)])
        wb.save(EXCEL_PATH)

        messagebox.showinfo("가입 완료", "회원 가입이 완료되었습니다.")
        join_window.destroy()
        login_frame.pack(pady=50)  # 로그인 창 다시 보여줌

    tk.Button(join_window, text="가입하기", command=save_user).pack(pady=10)

# ===== 로그인 검사 =====

def check_login(entry_id, entry_pw):
    user_id = entry_id.get().strip()
    password = entry_pw.get().strip()

    credentials = load_credentials(EXCEL_PATH)

    if user_id not in credentials:
        messagebox.showinfo("로그인 실패", "존재하지 않는 아이디입니다.")
    elif not check_password(password, credentials[user_id]):
        messagebox.showwarning("로그인 실패", "비밀번호가 틀렸습니다.")
    else:
        messagebox.showinfo("로그인 성공", f"{user_id}님 환영합니다!")

# ===== GUI =====

root = tk.Tk()
root.title("로그인")

default_window(root, 450, 650)

login_frame = tk.Frame(root)
login_frame.pack(pady=50)

tk.Label(login_frame, text="로그인", font=("Arial", 16)).pack(pady=10)

tk.Label(login_frame, text="ID").pack()
entry_id = tk.Entry(login_frame)
entry_id.pack()

tk.Label(login_frame, text="Password").pack()
entry_pw = tk.Entry(login_frame, show="*")
entry_pw.pack()

tk.Button(login_frame, text="로그인", command=lambda: check_login(entry_id, entry_pw)).pack(pady=10)
tk.Button(login_frame, text="회원가입", command=lambda: open_join_window(login_frame)).pack()

root.mainloop()
