import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
import bcrypt

# 페이지 이동을 위한 subprocess
import subprocess

EXCEL_PATH = "USER_DOCS.xlsx"

# 엑셀 파일에서 사용자 정보 불러오기
def load_credentials(EXCEL_PATH):
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # 사용자 ID와 비밀번호를 저장할 딕셔너리
    credentials = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        user_id, password = row
        credentials[str(user_id).strip()] = str(password).strip()
    return credentials

# 로그인 검사
def check_login():
    user_id = entry_id.get()
    password = entry_pw.get()

    credentials = load_credentials(EXCEL_PATH)

    if user_id not in credentials:
        messagebox.showinfo("로그인 실패", "존재하지 않는 아이디 입니다.")
    elif not check_password(password, credentials[user_id]):
        messagebox.showwarning("로그인 실패", "비밀번호가 틀렸습니다.")
    else :
        print("Login Successful!")
        messagebox.showinfo("로그인 성공", "환영합니다!")

# 비밀번호 확인
def check_password(plain_password, hashed_password):
    return bcrypt.checkpw(plain_password.encode(), hashed_password.encode())

def open_join_page():
    root.destroy()
    subprocess.run(["python3", "join_page.py"])

# GUI 구성
root = tk.Tk()
root.title("로그인")

root.attributes("-topmost", True)

window_width = 450
window_height = 650

# 모니터 정중앙에 배치
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

center_x = int((screen_width - window_width) / 2)
center_y = int((screen_height - window_height) / 2)

root.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")

# 로고 부분
logo_frame = tk.Frame(root, width=150, height=150, bg="lightblue")
logo_frame.pack(pady=(100,100))
logo_frame.pack_propagate(False)

# ID 입력
tk.Label(root, text="ID").pack(anchor='w', padx=70)
entry_id = tk.Entry(root)
entry_id.pack(fill='x', padx=70, pady=(0, 10))

# Password 입력
tk.Label(root, text="Password").pack(anchor='w', padx=70)
entry_pw = tk.Entry(root, show="*")
entry_pw.pack(fill='x', padx=70, pady=(0, 20))

# 버튼 담을 프레임
button_frame = tk.Frame(root)
button_frame.pack(pady=20)

# 로그인 버튼
login_button = tk.Button(button_frame, text="로그인", command=check_login, width=12, height=2)
login_button.grid(row=0, column=0, padx=20)

# 회원가입 버튼
join_button = tk.Button(button_frame, text="회원가입", command=open_join_page, width=12, height=2)
join_button.grid(row=0, column=1, padx=20)

root.mainloop()