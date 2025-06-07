import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from bcrypt_utils import hash_password

EXCEL_PATH = "USER_DOCS.xlsx"

def load_credentials(path):
    credentials = {}
    try:
        wb = load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            credentials[row[0]] = row[1]
    except FileNotFoundError:
        pass
    return credentials

def create_join_frame(root, switch_to_login):
    frame = tk.Frame(root)

    tk.Label(frame, text="회원가입", font=("Arial", 16)).pack(pady=10)

    tk.Label(frame, text="ID").pack()
    entry_id = tk.Entry(frame)
    entry_id.pack()

    tk.Label(frame, text="Password").pack()
    entry_pw = tk.Entry(frame, show="*")
    entry_pw.pack()

    def save_user():
        user_id = entry_id.get().strip()
        password = entry_pw.get().strip()

        if not user_id or not password:
            messagebox.showwarning("입력 오류", "ID와 Password 모두 입력하세요.")
            return

        credentials = load_credentials(EXCEL_PATH)

        if user_id in credentials:
            messagebox.showerror("중복 오류", "이미 존재하는 ID입니다.")
            return

        try:
            wb = load_workbook(EXCEL_PATH)
        except FileNotFoundError:
            wb = Workbook()
            wb.active.append(["ID", "Password"])

        ws = wb.active
        ws.append([user_id, hash_password(password)])
        wb.save(EXCEL_PATH)

        messagebox.showinfo("가입 완료", "회원 가입이 완료되었습니다.")
        switch_to_login()

    tk.Button(frame, text="가입하기", command=save_user).pack(pady=10)

    return frame
