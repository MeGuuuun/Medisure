import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook

EXCEL_PATH = "USER_DOCS.xlsx"

def save_user():
    user_id = entry_id.get().strip()
    password = entry_pw.get().strip()

    if not user_id or not password:
        messagebox.showwarning("입력 오류", "ID와 Password 모두 입력하세요.")
        return

    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active

        # 이미 존재하는 ID 체크
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id:
                messagebox.showerror("중복 오류", "이미 존재하는 ID입니다.")
                return

        # 새 사용자 추가
        ws.append([user_id, password])
        wb.save(EXCEL_PATH)

        messagebox.showinfo("가입 완료", "회원 가입이 완료 되었습니다.")
        root.destroy()

    except FileNotFoundError:
        print("오류")


# GUI
root = tk.Tk()
root.title("회원가입")

root.attributes("-topmost", True)

window_width = 450
window_height = 650

screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

center_x = int((screen_width - window_width) / 2)
center_y = int((screen_height - window_height) / 2)

root.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")

# 로고 부분
logo_frame = tk.Frame(root, width=150, height=150, bg="lightblue")
logo_frame.pack(pady=(100,100))
logo_frame.pack_propagate(False)

tk.Label(root, text="회원가입").pack(pady=(30,20))

# ID 입력받기
tk.Label(root, text="ID").pack(anchor='w', padx=70)
entry_id = tk.Entry(root)
entry_id.pack(fill='x', padx=70, pady=(0, 10))

tk.Label(root, text="Password").pack(anchor='w', padx=70)
entry_pw = tk.Entry(root, show="*")
entry_pw.pack(fill='x', padx=70, pady=(0, 20))

tk.Button(root, text="가입하기", command=save_user).pack(pady=10)

root.mainloop()