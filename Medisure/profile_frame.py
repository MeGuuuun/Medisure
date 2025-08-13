import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
from data_fetcher import fetch_pill_info
from PIL import Image, ImageTk

EXCEL_PATH = "USER_DOCS.xlsx"
IMAGE_PATH = "Medisure_logo.png"

# 유저 정보 불러오기
def load_user_info(user_id):
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id:
                wb.close()
                return row
        wb.close()
        return None
    except Exception as e:
        print("⚠️ 엑셀 오류:", e)
        return None

# 팝업 창 중앙 정렬
def center_window(win):
    win.update_idletasks()
    x = (win.winfo_screenwidth() - win.winfo_width()) // 2
    y = (win.winfo_screenheight() - win.winfo_height()) // 2
    win.geometry(f"+{x}+{y}")

# 약물 추가 함수
def add_selected_to_excel(selected, user_id):
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active

        id_col = 1
        pill_cols = list(range(6, 11))
        target_row = None

        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=id_col).value == user_id:
                target_row = row
                break

        if not target_row:
            print("⚠️ 사용자 정보를 찾을 수 없습니다.")
            wb.close()
            return

        inserted = False
        for col in pill_cols:
            if not ws.cell(row=target_row, column=col).value:
                ws.cell(row=target_row, column=col, value=selected)
                inserted = True
                break

        if not inserted:
            messagebox.showwarning("공간 부족", "더 이상 약물을 추가할 공간이 없습니다.")

        wb.save(EXCEL_PATH)
        wb.close()
    except Exception as e:
        print("⚠️ 약물 추가 오류:", e)
        messagebox.showerror("오류", "약물 추가 중 문제가 발생했습니다.")

# 약물 정보 불러오는 함수
def refresh_pill_list(user_id, target_frame):
    # 기존에 표시된 약물 리스트 삭제 (중복 방지)
    for widget in target_frame.winfo_children():
        widget.destroy()

    pill_list_label = tk.Label(target_frame,text="복용 중인 약들", bg='white')
    pill_list_label.pack(pady=10)

    user_info = load_user_info(user_id)
    user_pills = user_info[5:11]

    # 약물 삭제 함수
    def delete_pill(pill_name):
        try:
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active

            row_index = None
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if str(row[0]) == user_id:
                    row_index = i
                    break

            if row_index is None:
                wb.close()
                messagebox.showerror("삭제 오류", "사용자 정보를 찾을 수 없습니다.")
                return

            pills = [ws.cell(row=row_index, column=col).value for col in range(6, 11)]
            pills = [pill for pill in pills if pill != pill_name]
            pills += [None] * (5 - len(pills))  # None으로 채워 5개 유지

            for idx, pill in enumerate(pills):
                ws.cell(row=row_index, column=6 + idx).value = pill

            wb.save(EXCEL_PATH)
            wb.close()

            refresh_pill_list(user_id, target_frame)
        except Exception as e:
            print("⚠️ 약물 삭제 중 오류:", e)
            messagebox.showerror("오류", "약물 삭제 중 문제가 발생했습니다.")

    for pill in user_pills:
        if pill:
            row_frame = tk.Frame(target_frame, bg='white')
            row_frame.pack(fill='x', padx=20, pady=3)

            label = tk.Label(row_frame, text=f"- {pill}", anchor="w", bg='white')
            label.pack(side='left', fill='x', expand=True)

            del_btn = tk.Button(row_frame, text="❌", command=lambda p=pill: delete_pill(p),
                                bg='white', fg='black', relief='solid', bd=1, padx=5, pady=1)
            del_btn.pack(side='right', padx=(5, 0))

# profile 프레임 생성
def create_profile_frame(root,user_id, on_logout, switch_to_interaction):
    frame = tk.Frame(root)

    # ==== Frame 선언 ====
    logo_frame = tk.Frame(frame, height=150, width=150)
    logo_frame.pack(fill="x", pady=(30, 10))
    logo_frame.pack_propagate(False)

    logo_img_raw = Image.open(IMAGE_PATH)
    logo_img_resized = logo_img_raw.resize((150, 150), Image.Resampling.LANCZOS)
    logo_img = ImageTk.PhotoImage(logo_img_resized)

    logo_label = tk.Label(logo_frame, image=logo_img)
    logo_label.image = logo_img
    logo_label.pack()

    tk.Label(frame, text="나의 약 관리", font=("Arial", 16)).pack(pady=10)

    search_frame = tk.Frame(frame)
    search_frame.pack(fill="x", padx=10, pady=10)

    bottom_frame = tk.Frame(frame)
    bottom_frame.pack(fill="both", expand=True, padx=10, pady=10)
    refresh_pill_list(user_id, bottom_frame)

    # ==== 약물 검색 =====
    tk.Label(search_frame, text="약물명 입력", font=("Arial", 12)).pack()

    entry_btn_frame = tk.Frame(search_frame)
    entry_btn_frame.pack(pady=5)
    entry_btn_frame.pack_configure(padx=20)  # 좌우 20픽셀 여백 예시

    pill_entry = tk.Entry(entry_btn_frame, font=("Arial", 12), width=25, fg='grey', bg='lightgrey')
    pill_entry.pack(side="left")

    def search_pill():
        pill_name = pill_entry.get().strip()
        if not pill_name:
            messagebox.showwarning("입력 오류", "약물명을 입력하세요.")
            return

        try:
            results = fetch_pill_info(pill_name)
        except Exception as e:
            print("⚠️ 약물 검색 오류:", e)
            messagebox.showerror("검색 오류", "약물 정보를 불러오는 중 오류가 발생했습니다.")
            return

        if not results:
            messagebox.showinfo("검색 결과 없음", f"'{pill_name}'에 대한 검색 결과가 없습니다.")
            return

        # 검색 결과 팝업 창 생성
        popup = tk.Toplevel(search_frame)
        popup.geometry("300x150")
        popup.resizable(False, False)
        popup.transient(search_frame)
        popup.grab_set()
        center_window(popup)

        scrollbar = tk.Scrollbar(popup)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        listbox = tk.Listbox(popup, yscrollcommand=scrollbar.set, width=50, height=15)
        for name in results:
            listbox.insert(tk.END, name)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=listbox.yview)

        def on_select(event):
            if listbox.curselection():
                selected = listbox.get(listbox.curselection())
                try:
                    add_selected_to_excel(selected, user_id)
                    refresh_pill_list(user_id, bottom_frame)
                except Exception as e:
                    print("⚠️ 선택 후 오류:", e)
                    messagebox.showerror("오류", "약물을 추가하는 도중 문제가 발생했습니다.")
                    popup.destroy()
                    return

                index = listbox.curselection()[0]
                listbox.itemconfig(index, bg="lightgray")
                popup.after(500, popup.destroy)

        listbox.bind("<<ListboxSelect>>", on_select)


    search_btn = tk.Button(entry_btn_frame, text="약물 검색", font=("Arial", 12), command=search_pill)
    search_btn.pack(side="left", padx=(15, 0))

    # ==== 약물 정보 ====

    refresh_pill_list(user_id, bottom_frame)

    # === 약물 상호작용 ====

    action_frame = tk.Frame(frame)
    action_frame.pack(pady=20)

    tk.Button(action_frame, text="약물 상호작용 확인하기", font=("Arial", 12), command=switch_to_interaction).pack(pady=5)
    tk.Button(action_frame, text="로그아웃", font=("Arial", 12), command=on_logout).pack(pady=5)

    return frame